import pandas as pd
import numpy as np
import json
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from modules.database_queries import execute_query

# Diccionario de shades (incluye 30ml y 45ml como está en tu script)
shades = {
    'IT00000000000000000000000000000021': '30ml Colorant - Light Auburn',
    'IT00000000000000000000000000000020': '30ml Colorant - Dark Auburn',
    'IT00000000000000000000000000000019': '30ml Colorant - Lightest Blond',
    'IT00000000000000000000000000000018': '30ml Colorant - Warm - Light Blond',
    'IT00000000000000000000000000000017': '30ml Colorant - Cool - Light Blond',
    'IT00000000000000000000000000000016': '30ml Colorant - Warm-Medium Blond',
    'IT00000000000000000000000000000015': '30ml Colorant - Cool-Medium Blond',
    'IT00000000000000000000000000000014': '30ml Colorant - Warm-Dark Blond',
    'IT00000000000000000000000000000013': '30ml Colorant - Dark Blond',
    'IT00000000000000000000000000000012': '30ml Colorant: Light Brown',
    'IT00000000000000000000000000000011': '30ml Colorant - Medium-Light Brown',
    'IT00000000000000000000000000000002': '30ml Colorant - Medium Brown',
    'IT00000000000000000000000000000010': '30ml Colorant - Medium-Dark Brown',
    'IT00000000000000000000000000000009': '30ml Colorant - Dark Brown',
    'IT00000000000000000000000000000008': '30ml Colorant - Soft-Black',
    'IT00000000000000000000000000000007': '30ml Colorant - Black',
    'IT00000000000000000000000000000006': '30ml Colorant - Jet-Black',

    'IT00000000000000000000000000000061': '45ml Colorant - Light Auburn',
    'IT00000000000000000000000000000060': '45ml Colorant - Dark Auburn',
    'IT00000000000000000000000000000059': '45ml Colorant - Lightest Blond',
    'IT00000000000000000000000000000058': '45ml Colorant - Warm - Light Blond',
    'IT00000000000000000000000000000057': '45ml Colorant - Cool - Light Blond',
    'IT00000000000000000000000000000056': '45ml Colorant - Warm-Medium Blond',
    'IT00000000000000000000000000000055': '45ml Colorant - Cool-Medium Blond',
    'IT00000000000000000000000000000054': '45ml Colorant - Warm-Dark Blond',
    'IT00000000000000000000000000000053': '45ml Colorant - Dark Blond',
    'IT00000000000000000000000000000052': '45ml Colorant: Light Brown',
    'IT00000000000000000000000000000051': '45ml Colorant - Medium-Light Brown',
    'IT00000000000000000000000000000050': '45ml Colorant - Medium Brown',
    'IT00000000000000000000000000000049': '45ml Colorant - Medium-Dark Brown',
    'IT00000000000000000000000000000048': '45ml Colorant - Dark Brown',
    'IT00000000000000000000000000000047': '45ml Colorant - Soft-Black',
    'IT00000000000000000000000000000046': '45ml Colorant - Black',
    'IT00000000000000000000000000000045': '45ml Colorant - Jet-Black',
}

# -----------------------------
# NUEVO: extractor genérico para leer variables del diagnostic (var 40/36, etc.)
# -----------------------------
def extract_diagnostic_var(additional_fields, var_ids):
    """
    Busca en diagnostic el primer value cuyo 'variable' esté en var_ids (en orden).
    Soporta:
      - diagnostic = list[dict]
      - diagnostic = dict con keys numéricas -> dict
      - diagnostic = dict con "values": (list o dict)
    Retorna string (p.ej. "118") o None.
    """
    if not additional_fields or pd.isna(additional_fields):
        return None

    try:
        diagnostic_data = json.loads(additional_fields)

        # normalizar "values"
        if isinstance(diagnostic_data, dict) and "values" in diagnostic_data:
            values = diagnostic_data.get("values")
        else:
            values = diagnostic_data

        # armar iterador de items
        items = []
        if isinstance(values, list):
            items = values
        elif isinstance(values, dict):
            # si es dict tipo {"396654": {"value":"14","variable":"4"}} o similar
            items = list(values.values())
        else:
            items = []

        # var_ids en orden: primero 40, luego 36
        for wanted_var in var_ids:
            for item in items:
                if not isinstance(item, dict):
                    continue
                var = item.get("variable", item.get("var"))
                val = item.get("value", item.get("val"))
                if var is None:
                    continue
                try:
                    var_int = int(var)
                except (ValueError, TypeError):
                    continue
                if var_int == int(wanted_var):
                    if val is None:
                        return None
                    return str(val)

        return None

    except json.JSONDecodeError:
        return None
    except Exception as e:
        print(f"Error procesando additionalFields (extract_diagnostic_var): {e}")
        return None


def map_experience_with_color(exp_code):
    """
    Mapea exp_code (string) a etiqueta.
    """
    if exp_code is None or (isinstance(exp_code, float) and pd.isna(exp_code)):
        return "Unknown"

    code = str(exp_code).strip()

    if code in ("118", "112"):
        return "Currently Dyed"
    if code in ("119", "116"):
        return "I've colored"
    if code in ("120", "114"):
        return "Never colored"
    return "Unknown"


# -----------------------------
# EXISTENTE: etnias (13/14/15) extractor
# -----------------------------
def extract_diagnostic_values(additional_fields):
    """
    Extrae los valores de diagnóstico (13, 14, 15) del campo additionalFields
    Devuelve una lista con los valores encontrados (solo el primero si hay múltiples)
    """
    if not additional_fields or pd.isna(additional_fields):
        return []

    try:
        diagnostic_data = json.loads(additional_fields)
        values_found = []

        # Formato 1: Lista de diccionarios [{"value": 13, "variable": 4}, ...]
        if isinstance(diagnostic_data, list):
            for item in diagnostic_data:
                if isinstance(item, dict) and 'value' in item:
                    try:
                        value = int(item['value'])
                        if value in [13, 14, 15]:
                            values_found.append(value)
                    except (ValueError, TypeError):
                        continue

        # Formato 2: Diccionario {"396654": {"value": "14", "variable": "4"}}
        elif isinstance(diagnostic_data, dict):
            for _, value_dict in diagnostic_data.items():
                if isinstance(value_dict, dict) and 'value' in value_dict:
                    try:
                        value = int(value_dict['value'])
                        if value in [13, 14, 15]:
                            values_found.append(value)
                    except (ValueError, TypeError):
                        continue
                elif isinstance(value_dict, list):
                    for item in value_dict:
                        if isinstance(item, dict) and 'value' in item:
                            try:
                                value = int(item['value'])
                                if value in [13, 14, 15]:
                                    values_found.append(value)
                            except (ValueError, TypeError):
                                continue

        return values_found[:1] if values_found else []

    except json.JSONDecodeError:
        return []
    except Exception as e:
        print(f"Error procesando additionalFields: {e}")
        return []


def procesar_razon(razon):
    """
    Procesa la razón de cancelación para obtener más detalle en casos específicos
    (en tu script special_cases está vacío, así que se queda igual)
    """
    if pd.isna(razon):
        return "Sin razón especificada"

    razon_str = str(razon).strip()

    special_cases = []

    for case in special_cases:
        if razon_str.startswith(case):
            if razon_str.count('->') >= 2:
                parts = razon_str.split('->', 2)
                return f"{parts[0].strip()} -> {parts[1].strip()}"
            elif '->' in razon_str:
                return razon_str
            else:
                return razon_str

    if '->' in razon_str:
        return razon_str.split('->')[0].strip()
    else:
        return razon_str


def agregar_porcentaje(valor):
    if pd.isna(valor):
        return ""
    return f"{valor}%"


def obtener_suscripciones_activas(startDate, endDate):
    """
    Obtiene suscripciones (para tablas de distribución) por etnia, shade
    y AHORA también experience_with_color.
    """
    item_ids = list(shades.keys())
    item_ids_str = "', '".join(item_ids)

    query = f"""
    SELECT  
        sub.id,
        sub.additionalFields->>"$.diagnostic" AS additionalFields,
        subIt.itemId
    FROM prod_sales_and_subscriptions.subscriptions sub
    JOIN bi.fact_orders fo ON sub.id = fo.subscription_id
    JOIN prod_sales_and_subscriptions.subscription_items subIt on sub.id = subIt.subscriptionId
    WHERE fo.status NOT IN ('CANCELLED','PAYMENT_ERROR')
    AND fo.created_at BETWEEN '{startDate}' AND '{endDate}'
    AND subIt.itemId IN ('{item_ids_str}')
    GROUP BY sub.id;
    """

    df_suscripciones = execute_query(query)

    # Etnias
    df_suscripciones['diagnostic_values'] = df_suscripciones['additionalFields'].apply(extract_diagnostic_values)
    df_suscripciones['has_13'] = df_suscripciones['diagnostic_values'].apply(lambda x: 13 in x)
    df_suscripciones['has_14'] = df_suscripciones['diagnostic_values'].apply(lambda x: 14 in x)
    df_suscripciones['has_15'] = df_suscripciones['diagnostic_values'].apply(lambda x: 15 in x)

    # Experience (var 40 preferido; fallback var 36)
    df_suscripciones['exp_code'] = df_suscripciones['additionalFields'].apply(
        lambda x: extract_diagnostic_var(x, [40, 36])
    )
    df_suscripciones['experience_with_color'] = df_suscripciones['exp_code'].apply(map_experience_with_color)

    # Shades
    for item_id, shade_name in shades.items():
        df_suscripciones[f'shade_{item_id[-4:]}'] = df_suscripciones['itemId'] == item_id

    return df_suscripciones


def crear_tabla_etnias(df_suscripciones):
    total_suscripciones = len(df_suscripciones)

    tabla_etnias = pd.DataFrame({
        'Etnia': ['CAUCASIAN', 'AFRICAN', 'ASIAN', 'Sin diagnóstico'],
        'Total Suscripciones': [
            df_suscripciones['has_13'].sum(),
            df_suscripciones['has_14'].sum(),
            df_suscripciones['has_15'].sum(),
            total_suscripciones - (df_suscripciones['has_13'].sum() + df_suscripciones['has_14'].sum() + df_suscripciones['has_15'].sum())
        ]
    })

    if total_suscripciones > 0:
        tabla_etnias['Porcentaje'] = (tabla_etnias['Total Suscripciones'] / total_suscripciones * 100).round(2)
        tabla_etnias['Porcentaje'] = tabla_etnias['Porcentaje'].apply(agregar_porcentaje)
    else:
        tabla_etnias['Porcentaje'] = "0%"

    return tabla_etnias


def crear_tabla_shades(df_suscripciones):
    total_suscripciones = len(df_suscripciones)

    datos_shades = []
    for item_id, shade_name in shades.items():
        col_name = f'shade_{item_id[-4:]}'
        count = df_suscripciones[col_name].sum()
        datos_shades.append({
            'Shade': shade_name,
            'Total Suscripciones': count
        })

    tabla_shades = pd.DataFrame(datos_shades)

    if total_suscripciones > 0:
        tabla_shades['Porcentaje'] = (tabla_shades['Total Suscripciones'] / total_suscripciones * 100).round(2)
        tabla_shades['Porcentaje'] = tabla_shades['Porcentaje'].apply(agregar_porcentaje)
    else:
        tabla_shades['Porcentaje'] = "0%"

    tabla_shades = tabla_shades.sort_values('Total Suscripciones', ascending=False)

    return tabla_shades


# -----------------------------
# NUEVO: tabla de distribución de suscripciones activas por experience
# -----------------------------
def crear_tabla_experience(df_suscripciones):
    total = len(df_suscripciones)
    order = ["Currently Dyed", "I've colored", "Never colored", "Unknown"]

    counts = (
        df_suscripciones["experience_with_color"]
        .fillna("Unknown")
        .value_counts()
        .reindex(order, fill_value=0)
        .reset_index()
    )
    counts.columns = ["Experience with color", "Total Suscripciones"]

    if total > 0:
        counts["Porcentaje"] = (counts["Total Suscripciones"] / total * 100).round(2).apply(agregar_porcentaje)
    else:
        counts["Porcentaje"] = "0%"

    return counts


def analizar_cancelaciones_por_razon(df):
    df['razon_procesada'] = df['reason'].apply(procesar_razon)

    resultado = df.groupby('razon_procesada').agg(
        total_cancelaciones=('id', 'count'),
        suscripciones_unicas=('subscriptionId', 'nunique'),
        caucasian=('has_13', 'sum'),
        african=('has_14', 'sum'),
        asian=('has_15', 'sum')
    ).reset_index()

    total_cancelaciones = resultado['total_cancelaciones'].sum()

    if total_cancelaciones > 0:
        resultado['porcentaje_cancelaciones'] = (resultado['total_cancelaciones'] / total_cancelaciones * 100).round(2)
    else:
        resultado['porcentaje_cancelaciones'] = 0

    total_caucasian = resultado['caucasian'].sum()
    total_african = resultado['african'].sum()
    total_asian = resultado['asian'].sum()

    resultado['porcentaje_caucasian'] = (resultado['caucasian'] / total_caucasian * 100).round(2) if total_caucasian > 0 else 0
    resultado['porcentaje_african'] = (resultado['african'] / total_african * 100).round(2) if total_african > 0 else 0
    resultado['porcentaje_asian'] = (resultado['asian'] / total_asian * 100).round(2) if total_asian > 0 else 0

    for col in ['porcentaje_cancelaciones', 'porcentaje_caucasian', 'porcentaje_african', 'porcentaje_asian']:
        resultado[col] = resultado[col].apply(agregar_porcentaje)

    resultado = resultado.sort_values('total_cancelaciones', ascending=False)

    return resultado


def analizar_cancelaciones_por_razon_y_shade(df, filtro_etnia=None):
    if filtro_etnia == 13:
        df = df[df['has_13'] == True]
        nombre_etnia = "CAUCASIAN"
    elif filtro_etnia == 14:
        df = df[df['has_14'] == True]
        nombre_etnia = "AFRICAN"
    elif filtro_etnia == 15:
        df = df[df['has_15'] == True]
        nombre_etnia = "ASIAN"
    else:
        nombre_etnia = "TODAS"

    df['razon_procesada'] = df['reason'].apply(procesar_razon)

    for item_id, shade_name in shades.items():
        col_name = f"shade_{item_id[-4:]}"
        df[col_name] = df['itemIds'].str.contains(item_id, na=False)

    agg_dict = {
        'total_cancelaciones': ('id', 'count'),
        'suscripciones_unicas': ('subscriptionId', 'nunique')
    }

    for item_id in shades.keys():
        col_name = f"shade_{item_id[-4:]}"
        agg_dict[col_name] = (col_name, 'sum')

    resultado = df.groupby('razon_procesada').agg(**agg_dict).reset_index()

    total_cancelaciones = resultado['total_cancelaciones'].sum()

    if total_cancelaciones > 0:
        resultado['porcentaje_cancelaciones'] = (resultado['total_cancelaciones'] / total_cancelaciones * 100).round(2)
    else:
        resultado['porcentaje_cancelaciones'] = 0

    for item_id in shades.keys():
        col_name = f"shade_{item_id[-4:]}"
        total_shade = resultado[col_name].sum()
        if total_shade > 0:
            resultado[f'porcentaje_{col_name}'] = (resultado[col_name] / total_shade * 100).round(2)
        else:
            resultado[f'porcentaje_{col_name}'] = 0

    rename_dict = {}
    for item_id, shade_name in shades.items():
        col_name = f"shade_{item_id[-4:]}"
        rename_dict[col_name] = shade_name
        rename_dict[f'porcentaje_{col_name}'] = f'porcentaje_{shade_name}'

    resultado = resultado.rename(columns=rename_dict)

    columnas_porcentaje = ['porcentaje_cancelaciones'] + [f'porcentaje_{shade_name}' for shade_name in shades.values()]
    for col in columnas_porcentaje:
        if col in resultado.columns:
            resultado[col] = resultado[col].apply(agregar_porcentaje)

    resultado = resultado.sort_values('total_cancelaciones', ascending=False)

    return resultado, nombre_etnia


# -----------------------------
# NUEVO: análisis por razón + experience_with_color
# -----------------------------
def analizar_cancelaciones_por_razon_y_experience(df):
    df = df.copy()
    df["razon_procesada"] = df["reason"].apply(procesar_razon)

    # Normaliza experiencia
    df["experience_with_color"] = df["experience_with_color"].fillna("Unknown")

    # Booleans por categoría
    df["exp_currently_dyed"] = df["experience_with_color"] == "Currently Dyed"
    df["exp_ive_colored"] = df["experience_with_color"] == "I've colored"
    df["exp_never_colored"] = df["experience_with_color"] == "Never colored"
    df["exp_unknown"] = df["experience_with_color"] == "Unknown"

    resultado = df.groupby("razon_procesada").agg(
        total_cancelaciones=("id", "count"),
        suscripciones_unicas=("subscriptionId", "nunique"),
        currently_dyed=("exp_currently_dyed", "sum"),
        ive_colored=("exp_ive_colored", "sum"),
        never_colored=("exp_never_colored", "sum"),
        unknown=("exp_unknown", "sum"),
    ).reset_index()

    total_cancelaciones = resultado["total_cancelaciones"].sum()
    if total_cancelaciones > 0:
        resultado["porcentaje_cancelaciones"] = (resultado["total_cancelaciones"] / total_cancelaciones * 100).round(2)
    else:
        resultado["porcentaje_cancelaciones"] = 0

    # Porcentajes por cada categoría (sobre el total de la categoría)
    for col, pct_col in [
        ("currently_dyed", "porcentaje_currently_dyed"),
        ("ive_colored", "porcentaje_ive_colored"),
        ("never_colored", "porcentaje_never_colored"),
        ("unknown", "porcentaje_unknown"),
    ]:
        denom = resultado[col].sum()
        resultado[pct_col] = (resultado[col] / denom * 100).round(2) if denom > 0 else 0

    # Formato %
    for col in [
        "porcentaje_cancelaciones",
        "porcentaje_currently_dyed",
        "porcentaje_ive_colored",
        "porcentaje_never_colored",
        "porcentaje_unknown",
    ]:
        resultado[col] = resultado[col].apply(agregar_porcentaje)

    return resultado.sort_values("total_cancelaciones", ascending=False)


def ajustar_ancho_columnas(archivo_excel):
    try:
        wb = load_workbook(archivo_excel)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(archivo_excel)
        print(f"✓ Ancho de columnas ajustado para {archivo_excel}")

    except Exception as e:
        print(f"✗ Error ajustando el ancho de columnas para {archivo_excel}: {e}")


def main(startDate, endDate):
    # 1. Obtener datos de cancelaciones
    item_ids = list(shades.keys())
    item_ids_str = "', '".join(item_ids)

    query = f"""
    WITH cancels AS (
        SELECT
            fc.id,
            fc.subscriptionId,
            fc.reason,
            fc.createdAt
        FROM bi.fact_cancellations fc
        WHERE fc.createdAt BETWEEN '{startDate}' AND '{endDate}'
        ),

        last_order AS (
        SELECT *
        FROM (
            SELECT
            c.id AS cancel_id,
            fo.id AS salesOrderId,
            fo.order_number,
            fo.created_at,
            ROW_NUMBER() OVER (
                PARTITION BY c.id
                ORDER BY fo.created_at DESC
            ) AS rn
            FROM cancels c
            JOIN bi.fact_orders fo
            ON fo.subscription_id = c.subscriptionId
            AND fo.status NOT IN ('CANCELLED','PAYMENT_ERROR')
            AND fo.created_at < c.createdAt
        ) t
        WHERE t.rn = 1
        ),

        items AS (
        SELECT
            lo.cancel_id,
            GROUP_CONCAT(DISTINCT fso.itemId) AS itemIds
        FROM last_order lo
        JOIN bi.fact_sales_order_items fso
            ON fso.salesOrderId = lo.salesOrderId
        GROUP BY lo.cancel_id
        ),

        diagnostic AS (
        SELECT
            lo.cancel_id,
            COALESCE(
            soi.additionalFields->>"$.diagnostic",
            so.additionalFields->>"$.diagnostic"
            ) AS additionalFields
        FROM last_order lo
        JOIN prod_sales_and_subscriptions.sales_orders so
            ON so.id = lo.salesOrderId
        LEFT JOIN prod_sales_and_subscriptions.sales_order_items soi
            ON soi.salesOrderId = so.id
        AND soi.itemId LIKE "%0001004170%"
        )

        SELECT
        c.id,
        c.subscriptionId,
        c.reason,
        c.createdAt,
        i.itemIds,
        lo.order_number AS orderNumbers,
        d.additionalFields
        FROM cancels c
        JOIN last_order lo
        ON lo.cancel_id = c.id
        JOIN items i
        ON i.cancel_id = c.id
        LEFT JOIN diagnostic d
        ON d.cancel_id = c.id
        WHERE EXISTS (
        SELECT 1
        FROM bi.fact_sales_order_items fso2
        WHERE fso2.salesOrderId = lo.salesOrderId
            AND fso2.itemId IN ('{item_ids_str}')
        );
    """

    df = execute_query(query)

    # Etnias
    print("Procesando datos de diagnóstico (etnias)...")
    df['diagnostic_values'] = df['additionalFields'].apply(extract_diagnostic_values)
    df['has_13'] = df['diagnostic_values'].apply(lambda x: 13 in x)
    df['has_14'] = df['diagnostic_values'].apply(lambda x: 14 in x)
    df['has_15'] = df['diagnostic_values'].apply(lambda x: 15 in x)

    # NUEVO: Experience with color
    print("Procesando datos de diagnóstico (experience_with_color)...")
    df["exp_code"] = df["additionalFields"].apply(lambda x: extract_diagnostic_var(x, [40, 36]))
    df["experience_with_color"] = df["exp_code"].apply(map_experience_with_color)

    # 2. Suscripciones activas (para tablas)
    print("Obteniendo datos de suscripciones activas...")
    df_suscripciones = obtener_suscripciones_activas(startDate, endDate)

    print("Creando tablas de suscripciones...")
    tabla_etnias = crear_tabla_etnias(df_suscripciones)
    tabla_shades = crear_tabla_shades(df_suscripciones)
    tabla_experience = crear_tabla_experience(df_suscripciones)

    # 3. Tablas de cancelaciones
    print("Procesando datos de cancelaciones por razón (con etnias)...")
    df_por_razon = analizar_cancelaciones_por_razon(df)

    print("Procesando datos de cancelaciones por razón (con shades)...")
    df_por_razon_y_shade, _ = analizar_cancelaciones_por_razon_y_shade(df)

    print("Procesando datos de cancelaciones por razón (CAUCASIAN)...")
    df_caucasian, _ = analizar_cancelaciones_por_razon_y_shade(df, 13)

    print("Procesando datos de cancelaciones por razón (AFRICAN)...")
    df_african, _ = analizar_cancelaciones_por_razon_y_shade(df, 14)

    print("Procesando datos de cancelaciones por razón (ASIAN)...")
    df_asian, _ = analizar_cancelaciones_por_razon_y_shade(df, 15)

    # NUEVO: Por razón (Experience)
    print("Procesando datos de cancelaciones por razón (Experience)...")
    df_por_razon_experience = analizar_cancelaciones_por_razon_y_experience(df)

    # 4. Guardar Excel
    nombre_archivo = f"analisis_cancelaciones_{startDate}_to_{endDate}.xlsx"

    with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
        # Hoja 1: Por Razon (Etnias)
        df_por_razon.to_excel(writer, sheet_name='Por Razon (Etnias)', index=False, startrow=0)
        tabla_etnias.to_excel(writer, sheet_name='Por Razon (Etnias)', index=False, startrow=len(df_por_razon) + 3)

        # Hoja 2: Por Razon (Shades)
        start_row = 0
        df_por_razon_y_shade.to_excel(writer, sheet_name='Por Razon (Shades)', index=False, startrow=start_row)
        start_row += len(df_por_razon_y_shade) + 3

        writer.sheets['Por Razon (Shades)'].cell(row=start_row, column=1, value="CAUCASIAN (13)")
        start_row += 1
        df_caucasian.to_excel(writer, sheet_name='Por Razon (Shades)', index=False, startrow=start_row)
        start_row += len(df_caucasian) + 3

        writer.sheets['Por Razon (Shades)'].cell(row=start_row, column=1, value="AFRICAN (14)")
        start_row += 1
        df_african.to_excel(writer, sheet_name='Por Razon (Shades)', index=False, startrow=start_row)
        start_row += len(df_african) + 3

        writer.sheets['Por Razon (Shades)'].cell(row=start_row, column=1, value="ASIAN (15)")
        start_row += 1
        df_asian.to_excel(writer, sheet_name='Por Razon (Shades)', index=False, startrow=start_row)
        start_row += len(df_asian) + 3

        writer.sheets['Por Razon (Shades)'].cell(row=start_row, column=1, value="Suscripciones Activas por Shade")
        start_row += 1
        tabla_shades.to_excel(writer, sheet_name='Por Razon (Shades)', index=False, startrow=start_row)

        # Hoja 3: Por Razon (Experience)
        df_por_razon_experience.to_excel(writer, sheet_name='Por Razon (Experience)', index=False, startrow=0)
        tabla_experience.to_excel(
            writer,
            sheet_name='Por Razon (Experience)',
            index=False,
            startrow=len(df_por_razon_experience) + 3
        )

    # 5. Ajustar columnas
    ajustar_ancho_columnas(nombre_archivo)

    print(f"\n✅ Análisis completado. Archivo guardado como: {nombre_archivo}")
    print(f"Total de cancelaciones: {df_por_razon['total_cancelaciones'].sum()}")
    print(f"Total de suscripciones únicas canceladas: {df_por_razon['suscripciones_unicas'].sum()}")
    print(f"Total de suscripciones activas: {len(df_suscripciones)}")

    # Stats etnias
    print(f"Cancelaciones CAUCASIAN (13): {df['has_13'].sum()}")
    print(f"Cancelaciones AFRICAN (14): {df['has_14'].sum()}")
    print(f"Cancelaciones ASIAN (15): {df['has_15'].sum()}")
    print(f"Cancelaciones sin diagnóstico etnia: {len(df) - (df['has_13'].sum() + df['has_14'].sum() + df['has_15'].sum())}")

    # Stats experience
    print("\nDistribución cancelaciones por experience_with_color:")
    print(df["experience_with_color"].value_counts(dropna=False))

    print("\n📋 Top 5 razones de cancelación (Etnias):")
    print(df_por_razon[['razon_procesada', 'total_cancelaciones', 'porcentaje_cancelaciones']].head())

    print("\n📋 Distribución de suscripciones activas por etnia:")
    print(tabla_etnias)

    print("\n📋 Distribución de suscripciones activas por experience_with_color:")
    print(tabla_experience)


if __name__ == "__main__":
    main('2021-01-01', '2021-04-01')
    # main('2021-03-01', '2021-07-01')
    # main('2021-07-01', '2021-10-01')
    # main('2021-10-01', '2022-01-01')

    # main('2022-01-01', '2022-04-01')
    # main('2022-03-01', '2022-07-01')
    # main('2022-07-01', '2022-10-01')
    # main('2022-10-01', '2023-01-01')

    # main('2023-01-01', '2023-04-01')
    # main('2023-03-01', '2023-07-01')
    # main('2023-07-01', '2023-10-01')
    # main('2023-10-01', '2024-01-01')
    
    # main('2024-01-01', '2024-04-01')
    # main('2024-03-01', '2024-07-01')
    # main('2024-07-01', '2024-10-01')
    # main('2024-10-01', '2025-01-01')

    # main('2025-01-01', '2025-04-01')
    # main('2025-03-01', '2025-07-01')
    # main('2025-07-01', '2025-10-01')
    # main('2025-10-01', '2026-01-01')