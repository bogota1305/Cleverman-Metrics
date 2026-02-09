import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from modules.database_queries import execute_query


# Diccionario completo de colorants (30ml + 45ml)
shades = {
    # 30ml
    'IT00000000000000000000000000000021': '30ml Colorant - Light Auburn',
    'IT00000000000000000000000000000020': '30ml Colorant - Dark Auburn',
    'IT00000000000000000000000000000019': '30ml Colorant - Lightest Blond',
    'IT00000000000000000000000000000018': '30ml Colorant - Warm - Light Blond',
    'IT00000000000000000000000000000017': '30ml Colorant - Cool - Light Blond',
    'IT00000000000000000000000000000016': '30ml Colorant - Warm-Medium Blond',
    'IT00000000000000000000000000000015': '30ml Colorant - Cool-Medium Blond',
    'IT00000000000000000000000000000014': '30ml Colorant - Warm-Dark Blond',
    'IT00000000000000000000000000000013': '30ml Colorant - Dark Blond',
    'IT00000000000000000000000000000012': '30ml Colorant: Light Brown',
    'IT00000000000000000000000000000011': '30ml Colorant - Medium-Light Brown',
    'IT00000000000000000000000000000002': '30ml Colorant - Medium Brown',
    'IT00000000000000000000000000000010': '30ml Colorant - Medium-Dark Brown',
    'IT00000000000000000000000000000009': '30ml Colorant - Dark Brown',
    'IT00000000000000000000000000000008': '30ml Colorant - Soft-Black',
    'IT00000000000000000000000000000007': '30ml Colorant - Black',
    'IT00000000000000000000000000000006': '30ml Colorant - Jet-Black',

    # 45ml
    'IT00000000000000000000000000000061': '45ml Colorant - Light Auburn',
    'IT00000000000000000000000000000060': '45ml Colorant - Dark Auburn',
    'IT00000000000000000000000000000059': '45ml Colorant - Lightest Blond',
    'IT00000000000000000000000000000058': '45ml Colorant - Warm - Light Blond',
    'IT00000000000000000000000000000057': '45ml Colorant - Cool - Light Blond',
    'IT00000000000000000000000000000056': '45ml Colorant - Warm-Medium Blond',
    'IT00000000000000000000000000000055': '45ml Colorant - Cool-Medium Blond',
    'IT00000000000000000000000000000054': '45ml Colorant - Warm-Dark Blond',
    'IT00000000000000000000000000000053': '45ml Colorant - Dark Blond',
    'IT00000000000000000000000000000052': '45ml Colorant: Light Brown',
    'IT00000000000000000000000000000051': '45ml Colorant - Medium-Light Brown',
    'IT00000000000000000000000000000050': '45ml Colorant - Medium Brown',
    'IT00000000000000000000000000000049': '45ml Colorant - Medium-Dark Brown',
    'IT00000000000000000000000000000048': '45ml Colorant - Dark Brown',
    'IT00000000000000000000000000000047': '45ml Colorant - Soft-Black',
    'IT00000000000000000000000000000046': '45ml Colorant - Black',
    'IT00000000000000000000000000000045': '45ml Colorant - Jet-Black',
}


def procesar_razon(razon):
    if pd.isna(razon) or str(razon).strip() == "":
        return "Sin razón especificada"
    razon_str = str(razon).strip()
    if '->' in razon_str:
        return razon_str.split('->')[0].strip()
    return razon_str


def agregar_porcentaje(valor):
    if pd.isna(valor):
        return ""
    return f"{valor}%"


def ajustar_ancho_columnas(archivo_excel):
    try:
        wb = load_workbook(archivo_excel)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        wb.save(archivo_excel)
        print(f"✓ Ancho de columnas ajustado para {archivo_excel}")
    except Exception as e:
        print(f"✗ Error ajustando el ancho de columnas para {archivo_excel}: {e}")


def analizar_cancelaciones_por_razon(df_cancel_unico):
    """
    df_cancel_unico: 1 fila por cancelación (no por item)
    """
    df = df_cancel_unico.copy()
    df['razon_procesada'] = df['reason'].apply(procesar_razon)

    resultado = df.groupby('razon_procesada').agg(
        total_cancelaciones=('id', 'count'),
        suscripciones_unicas=('subscriptionId', 'nunique'),
        caucasian=('has_13', 'sum'),
        african=('has_14', 'sum'),
        asian=('has_15', 'sum'),
    ).reset_index()

    total_cancelaciones = int(resultado['total_cancelaciones'].sum())

    if total_cancelaciones > 0:
        resultado['porcentaje_cancelaciones'] = (resultado['total_cancelaciones'] / total_cancelaciones * 100).round(2)
    else:
        resultado['porcentaje_cancelaciones'] = 0

    total_caucasian = int(resultado['caucasian'].sum())
    total_african = int(resultado['african'].sum())
    total_asian = int(resultado['asian'].sum())

    resultado['porcentaje_caucasian'] = (resultado['caucasian'] / total_caucasian * 100).round(2) if total_caucasian else 0
    resultado['porcentaje_african'] = (resultado['african'] / total_african * 100).round(2) if total_african else 0
    resultado['porcentaje_asian'] = (resultado['asian'] / total_asian * 100).round(2) if total_asian else 0

    for col in ['porcentaje_cancelaciones', 'porcentaje_caucasian', 'porcentaje_african', 'porcentaje_asian']:
        resultado[col] = resultado[col].apply(agregar_porcentaje)

    return resultado.sort_values('total_cancelaciones', ascending=False)


def _build_cancel_level_table(df_items):
    """
    Convierte df a nivel cancelación:
    - key = subscriptionId + cancelAt
    - reason: tomamos la primera (debería ser consistente)
    - etnicity: primera
    - salesOrderNumberBeforeCancel / orderAtBeforeCancel: primera
    """
    cancel_key = ['subscriptionId', 'cancelAt']

    df_cancel = (
        df_items.sort_values(cancel_key)
        .groupby(cancel_key, as_index=False)
        .agg(
            reason=('reason', 'first'),
            etnicity=('etnicity', 'first'),
            salesOrderNumberBeforeCancel=('salesOrderNumberBeforeCancel', 'first'),
            orderAtBeforeCancel=('orderAtBeforeCancel', 'first'),
        )
    )

    # id artificial para conteos
    df_cancel['id'] = np.arange(1, len(df_cancel) + 1)

    # has_13/14/15 desde etnicity (ahora es numérico 13/14/15 o NULL)
    df_cancel['has_13'] = df_cancel['etnicity'].eq(13)
    df_cancel['has_14'] = df_cancel['etnicity'].eq(14)
    df_cancel['has_15'] = df_cancel['etnicity'].eq(15)

    return df_cancel


def analizar_cancelaciones_por_razon_y_shade(df_items, filtro_etnia=None):
    """
    df_items viene a nivel item (una fila por itemId del pedido antes de la cancelación).
    Para no inflar:
    - Creamos flags por cancelación y shade usando groupby + max().
    """
    cancel_key = ['subscriptionId', 'cancelAt']

    df = df_items.copy()

    # flags etnia a nivel fila (se mantienen para el groupby)
    df['has_13'] = df['etnicity'].eq(13)
    df['has_14'] = df['etnicity'].eq(14)
    df['has_15'] = df['etnicity'].eq(15)

    # filtrar por etnia si aplica
    if filtro_etnia == 13:
        df = df[df['has_13'] == True]
        nombre_etnia = "CAUCASIAN"
    elif filtro_etnia == 14:
        df = df[df['has_14'] == True]
        nombre_etnia = "AFRICAN"
    elif filtro_etnia == 15:
        df = df[df['has_15'] == True]
        nombre_etnia = "ASIAN"
    else:
        nombre_etnia = "TODAS"

    df['razon_procesada'] = df['reason'].apply(procesar_razon)

    # Crear flags por shade a nivel fila (item)
    for item_id in shades.keys():
        col = f"shade_{item_id[-4:]}"
        df[col] = (df['itemId'] == item_id)

    # Ahora colapsar a nivel cancelación (subscriptionId+cancelAt) con max()
    # (si el pedido tenía el shade, queda True)
    agg_cols = {'razon_procesada': ('razon_procesada', 'first')}
    for item_id in shades.keys():
        col = f"shade_{item_id[-4:]}"
        agg_cols[col] = (col, 'max')

    df_cancel_shade = (
        df.groupby(cancel_key, as_index=False)
          .agg(**agg_cols)
    )

    # id artificial de cancelación (para conteo)
    df_cancel_shade['id'] = np.arange(1, len(df_cancel_shade) + 1)

    # Agrupar por razón (ya a nivel cancelación)
    agg_dict = {
        'total_cancelaciones': ('id', 'count'),
        'suscripciones_unicas': ('subscriptionId', 'nunique'),
    }
    for item_id in shades.keys():
        col = f"shade_{item_id[-4:]}"
        agg_dict[col] = (col, 'sum')

    resultado = df_cancel_shade.groupby('razon_procesada').agg(**agg_dict).reset_index()

    # % total cancelaciones
    total_cancel = int(resultado['total_cancelaciones'].sum())
    resultado['porcentaje_cancelaciones'] = (resultado['total_cancelaciones'] / total_cancel * 100).round(2) if total_cancel else 0

    # % por shade
    for item_id in shades.keys():
        col = f"shade_{item_id[-4:]}"
        total_shade = int(resultado[col].sum())
        resultado[f'porcentaje_{col}'] = (resultado[col] / total_shade * 100).round(2) if total_shade else 0

    # Renombrar columnas a nombres legibles
    rename_dict = {}
    for item_id, shade_name in shades.items():
        base = f"shade_{item_id[-4:]}"
        rename_dict[base] = shade_name
        rename_dict[f'porcentaje_{base}'] = f'porcentaje_{shade_name}'
    resultado = resultado.rename(columns=rename_dict)

    # Agregar símbolo % a columnas porcentuales
    pct_cols = ['porcentaje_cancelaciones'] + [f'porcentaje_{name}' for name in shades.values()]
    for col in pct_cols:
        if col in resultado.columns:
            resultado[col] = resultado[col].apply(agregar_porcentaje)

    return resultado.sort_values('total_cancelaciones', ascending=False), nombre_etnia


def main(startDate, endDate, kitType):
    
    idType = 'IG00000000000000000000000000000028'
    if kitType == 'beard':
        idType = 'IG00000000000000000000000000000029'

    query = f"""
WITH cancel_with_last_order AS (
  SELECT
    ca.subscriptionId,
    ca.reason,
    ca.createdAt AS cancelAt,
    MAX(fo.created_at) AS lastOrderAt
  FROM prod_sales_and_subscriptions.cancellations ca
  LEFT JOIN bi.fact_orders fo
    ON fo.subscription_id = ca.subscriptionId
   AND fo.created_at < ca.createdAt
  WHERE ca.createdAt >= '{startDate}'
    AND ca.createdAt <  '{endDate}'
  GROUP BY
    ca.subscriptionId,
    ca.reason,
    ca.createdAt
),
dedup_items AS (
  SELECT DISTINCT
    soi.salesOrderId,
    soi.itemId
  FROM bi.fact_sales_order_items soi
)
SELECT
  c.subscriptionId,
  TRIM(SUBSTRING_INDEX(c.reason, '->', 1)) AS reason,
  c.cancelAt,

  fo.order_number AS salesOrderNumberBeforeCancel,
  fo.created_at   AS orderAtBeforeCancel,

  (
    SELECT
      CASE
        WHEN SUM(v = '14') > 0 THEN 14
        WHEN SUM(v = '15') > 0 THEN 15
        WHEN SUM(v = '13') > 0 THEN 13
        ELSE NULL
      END
    FROM (
      SELECT jt.v
      FROM JSON_TABLE(
        CASE
          WHEN JSON_TYPE(sub.additionalFields->'$.diagnostic.values') = 'ARRAY'
          THEN sub.additionalFields->'$.diagnostic.values'
          ELSE JSON_ARRAY()
        END,
        '$[*]' COLUMNS ( v VARCHAR(50) PATH '$.value' )
      ) jt

      UNION ALL

      SELECT jt2.v
      FROM JSON_TABLE(
        CASE
          WHEN JSON_TYPE(sub.additionalFields->'$.diagnostic.values') = 'OBJECT'
          THEN sub.additionalFields->'$.diagnostic.values'
          ELSE JSON_OBJECT()
        END,
        '$.*' COLUMNS ( v VARCHAR(50) PATH '$.value' )
      ) jt2
    ) x
  ) AS etnicity,

  di.itemId
FROM cancel_with_last_order c
LEFT JOIN bi.fact_orders fo
  ON fo.subscription_id = c.subscriptionId
 AND fo.created_at     = c.lastOrderAt
LEFT JOIN prod_sales_and_subscriptions.subscriptions sub
  ON sub.id = c.subscriptionId
JOIN dedup_items di
  ON di.salesOrderId = fo.id

WHERE c.lastOrderAt IS NOT NULL
  AND EXISTS (
    SELECT 1
    FROM bi.fact_sales_order_items oi
    WHERE oi.salesOrderId = fo.id
      AND oi.category = '{idType}'
  )
  AND di.itemId IN (
    # 'IT00000000000000000000000000000021',
    # 'IT00000000000000000000000000000020',
    # 'IT00000000000000000000000000000019',
    # 'IT00000000000000000000000000000018',
    # 'IT00000000000000000000000000000017',
    # 'IT00000000000000000000000000000016',
    # 'IT00000000000000000000000000000015',
    # 'IT00000000000000000000000000000014',
    # 'IT00000000000000000000000000000013',
    # 'IT00000000000000000000000000000012',
    # 'IT00000000000000000000000000000011',
    # 'IT00000000000000000000000000000002',
    # 'IT00000000000000000000000000000010',
    # 'IT00000000000000000000000000000009',
    # 'IT00000000000000000000000000000008'
    # 'IT00000000000000000000000000000007',
    # 'IT00000000000000000000000000000006',
    # 'IT00000000000000000000000000000061',
    # 'IT00000000000000000000000000000060',
    # 'IT00000000000000000000000000000059',
    # 'IT00000000000000000000000000000058',
    # 'IT00000000000000000000000000000057',
    # 'IT00000000000000000000000000000056',
    # 'IT00000000000000000000000000000055',
    # 'IT00000000000000000000000000000054',
    # 'IT00000000000000000000000000000053',
    # 'IT00000000000000000000000000000052',
    # 'IT00000000000000000000000000000051',
    # 'IT00000000000000000000000000000050',
    # 'IT00000000000000000000000000000049',
    # 'IT00000000000000000000000000000048',
    # 'IT00000000000000000000000000000047',
    # 'IT00000000000000000000000000000046',
    # 'IT00000000000000000000000000000045'
  );
"""

    df_items = execute_query(query)

    if df_items is None or df_items.empty:
        print("No se encontraron cancelaciones en el rango.")
        return

    # Asegurar tipos útiles
    df_items = df_items.copy()
    df_items['etnicity'] = pd.to_numeric(df_items['etnicity'], errors='coerce')  # 13/14/15 o NaN

    # 1) Tabla cancelación única (para "Por Razon (Etnias)")
    df_cancel_unico = _build_cancel_level_table(df_items)

    # 2) Tablas de salida
    print("Procesando datos de cancelaciones por razón (con etnias)...")
    df_por_razon = analizar_cancelaciones_por_razon(df_cancel_unico)

    print("Procesando datos de cancelaciones por razón (con shades)...")
    df_por_razon_y_shade, _ = analizar_cancelaciones_por_razon_y_shade(df_items)

    print("Procesando datos de cancelaciones por razón (CAUCASIAN)...")
    df_caucasian, _ = analizar_cancelaciones_por_razon_y_shade(df_items, 13)

    print("Procesando datos de cancelaciones por razón (AFRICAN)...")
    df_african, _ = analizar_cancelaciones_por_razon_y_shade(df_items, 14)

    print("Procesando datos de cancelaciones por razón (ASIAN)...")
    df_asian, _ = analizar_cancelaciones_por_razon_y_shade(df_items, 15)

    # 3) Excel (mismo layout general que antes, sin tablas de subs activas)
    nombre_archivo = f"analisis_cancelaciones_{startDate}_to_{endDate}_{kitType}_softBlack30ml.xlsx"

    with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
        df_por_razon.to_excel(writer, sheet_name='Por Razon (Etnias)', index=False, startrow=0)

        start_row = 0
        df_por_razon_y_shade.to_excel(writer, sheet_name='Por Razon (Shades)', index=False, startrow=start_row)
        start_row += len(df_por_razon_y_shade) + 3

        writer.sheets['Por Razon (Shades)'].cell(row=start_row, column=1, value="CAUCASIAN (13)")
        start_row += 1
        df_caucasian.to_excel(writer, sheet_name='Por Razon (Shades)', index=False, startrow=start_row)
        start_row += len(df_caucasian) + 3

        writer.sheets['Por Razon (Shades)'].cell(row=start_row, column=1, value="AFRICAN (14)")
        start_row += 1
        df_african.to_excel(writer, sheet_name='Por Razon (Shades)', index=False, startrow=start_row)
        start_row += len(df_african) + 3

        writer.sheets['Por Razon (Shades)'].cell(row=start_row, column=1, value="ASIAN (15)")
        start_row += 1
        df_asian.to_excel(writer, sheet_name='Por Razon (Shades)', index=False, startrow=start_row)

    ajustar_ancho_columnas(nombre_archivo)

    # 4) prints
    total_cancel = int(df_por_razon['total_cancelaciones'].sum())
    print(f"\n✅ Análisis completado. Archivo guardado como: {nombre_archivo}")
    print(f"Total de cancelaciones (únicas): {total_cancel}")
    print(f"Total de suscripciones únicas canceladas: {int(df_cancel_unico['subscriptionId'].nunique())}")

    print(f"Cancelaciones CAUCASIAN (13): {int(df_cancel_unico['has_13'].sum())}")
    print(f"Cancelaciones AFRICAN (14): {int(df_cancel_unico['has_14'].sum())}")
    print(f"Cancelaciones ASIAN (15): {int(df_cancel_unico['has_15'].sum())}")
    print(f"Cancelaciones sin diagnóstico: {int((~(df_cancel_unico['has_13'] | df_cancel_unico['has_14'] | df_cancel_unico['has_15'])).sum())}")

    print("\n📋 Top 5 razones de cancelación (Etnias):")
    print(df_por_razon[['razon_procesada', 'total_cancelaciones', 'porcentaje_cancelaciones']].head())


if __name__ == "__main__":
    main('2025-05-10', '2025-08-05', 'hair')
    main('2025-08-05', '2025-11-30', 'hair')
    # main('2025-07-01', '2025-10-01', 'hair')
    # main('2025-10-01', '2026-01-01', 'hair')

    # main('2025-01-01', '2025-04-01', 'beard')
    # main('2025-04-01', '2025-07-01', 'beard')
    # main('2025-07-01', '2025-10-01', 'beard')
    # main('2025-10-01', '2026-01-01', 'beard')
