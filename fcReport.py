
import pandas as pd
from pandas.api.types import is_datetime64_any_dtype
from openpyxl import load_workbook
from modules.database_queries import execute_query


fc_query = """
WITH renewals_with_number AS (
    SELECT
        r.id AS renewal_id,
        r.subscriptionId,
        r.createdAt AS renewal_createdAt,
        ROW_NUMBER() OVER (
            PARTITION BY r.subscriptionId 
            ORDER BY r.createdAt ASC
        ) AS renewal_number
    FROM prod_sales_and_subscriptions.renewals r
    WHERE r.status = 'SUCCESS'
)
SELECT
    DATE(CONVERT_TZ(v.first_date_sms_renewal_true, 'UTC', 'America/Los_Angeles')) AS full_control_starting_date,
    DATE(CONVERT_TZ(rw.renewal_createdAt, 'UTC', 'America/Los_Angeles')) AS renewal_date,
    DATE(CONVERT_TZ(v.last_date_sms_renewal_true, 'UTC', 'America/Los_Angeles')) AS full_control_ending_date,
    v.subscription_id,
    CASE
        WHEN rw.renewal_createdAt IS NULL THEN 0
        ELSE ROW_NUMBER() OVER (
            PARTITION BY v.subscription_id
            ORDER BY rw.renewal_createdAt ASC
        )
    END AS renewal_number,
    s.additionalFields->>"$.sms_renewal" AS full_control_active,
    s.status AS subscription_status
FROM prod_sales_and_subscriptions.first_sms_renewal_versions v
JOIN prod_sales_and_subscriptions.subscriptions s 
    ON s.id = v.subscription_id
LEFT JOIN renewals_with_number rw
    ON v.subscription_id = rw.subscriptionId
    AND DATE(rw.renewal_createdAt) >= DATE(v.first_date_sms_renewal_true) 
    AND (v.last_date_sms_renewal_true IS NULL OR DATE(rw.renewal_createdAt) <= DATE(v.last_date_sms_renewal_true))
WHERE 
    v.first_date_sms_renewal_true > '2023-01-01 08:00:00'
ORDER BY 
    v.subscription_id,
    rw.renewal_createdAt;
"""


MONTH_MAP = {
    1: "ene",
    2: "feb",
    3: "mar",
    4: "abr",
    5: "may",
    6: "jun",
    7: "jul",
    8: "ago",
    9: "sept",
    10: "oct",
    11: "nov",
    12: "dic",
}

FULL_CONTROL_SHEET_NAME = "Full control"
TEMPLATE_MONTH_HEADER_ROW = 2
TEMPLATE_FIRST_MONTH_COL = 2   # B
TEMPLATE_LAST_MONTH_COL = 26   # Z


def _normalize_date_column(df: pd.DataFrame, column_name: str) -> pd.Series:
    col = df[column_name]

    if is_datetime64_any_dtype(col):
        return pd.to_datetime(col, errors="coerce")

    parsed = pd.to_datetime(col, errors="coerce")

    numeric_mask = pd.to_numeric(col, errors="coerce").notna() & parsed.isna()
    if numeric_mask.any():
        serial_values = pd.to_numeric(col[numeric_mask], errors="coerce")
        parsed.loc[numeric_mask] = pd.to_datetime(
            serial_values,
            unit="D",
            origin="1899-12-30",
            errors="coerce"
        )

    return parsed


def _validate_required_columns(df: pd.DataFrame, required_cols: list[str], context: str = "") -> None:
    missing = [col for col in required_cols if col not in df.columns]
    if missing:
        prefix = f"{context}: " if context else ""
        raise ValueError(f"{prefix}faltan columnas requeridas: {missing}")


def _prepare_date_parts(df: pd.DataFrame, date_column: str) -> pd.DataFrame:
    data = df.copy()

    _validate_required_columns(data, [date_column], context="_prepare_date_parts")

    if not is_datetime64_any_dtype(data[date_column]):
        data[date_column] = _normalize_date_column(data, date_column)

    data = data[data[date_column].notna()].copy()
    data["year"] = data[date_column].dt.year
    data["month_number"] = data[date_column].dt.month
    data["month"] = data["month_number"].map(MONTH_MAP)

    return data


def _apply_single_filter(series: pd.Series, condition) -> pd.Series:
    """
    Devuelve una máscara booleana.

    Soporta:
    - valor simple: 1
    - {"op": "notna"}
    - {"op": "isna"}
    - {"op": "in", "values": [...]}
    - {"op": "eq", "value": x}
    - {"op": "ne", "value": x}
    - {"op": "gt", "value": x}
    - {"op": "gte", "value": x}
    - {"op": "lt", "value": x}
    - {"op": "lte", "value": x}
    """
    if not isinstance(condition, dict):
        numeric_series = pd.to_numeric(series, errors="coerce")
        if numeric_series.notna().any():
            return numeric_series == condition
        return series == condition

    op = condition.get("op")

    if op == "notna":
        return series.notna()

    if op == "isna":
        return series.isna()

    if op == "in":
        return series.isin(condition.get("values", []))

    value = condition.get("value")
    numeric_series = pd.to_numeric(series, errors="coerce")
    s = numeric_series if numeric_series.notna().any() else series

    if op == "eq":
        return s == value
    if op == "ne":
        return s != value
    if op == "gt":
        return s > value
    if op == "gte":
        return s >= value
    if op == "lt":
        return s < value
    if op == "lte":
        return s <= value

    raise ValueError(f"Operador de filtro no soportado: {op}")


def _apply_filters(df: pd.DataFrame, filters: dict | None = None) -> pd.DataFrame:
    data = df.copy()

    if not filters:
        return data

    for col, condition in filters.items():
        mask = _apply_single_filter(data[col], condition)
        data = data[mask].copy()

    return data


def _build_monthly_count_table(
    df: pd.DataFrame,
    date_column: str,
    value_column: str,
    output_column_name: str,
    filters: dict | None = None,
) -> pd.DataFrame:
    data = df.copy()

    required_cols = [date_column, value_column]
    if filters:
        required_cols.extend(filters.keys())

    _validate_required_columns(data, required_cols, context="_build_monthly_count_table")

    data = _prepare_date_parts(data, date_column)
    data = _apply_filters(data, filters)
    data = data[data[value_column].notna()].copy()

    result = (
        data.groupby(["year", "month_number", "month"], as_index=False)
        .agg(**{output_column_name: (value_column, "count")})
        .sort_values(["year", "month_number"])
        .reset_index(drop=True)
    )

    return result


def _build_monthly_status_table(
    df: pd.DataFrame,
    date_column: str,
    value_column: str,
    status_column: str,
    statuses: list[str],
    filters: dict | None = None,
) -> pd.DataFrame:
    data = df.copy()

    required_cols = [date_column, value_column, status_column]
    if filters:
        required_cols.extend(filters.keys())

    _validate_required_columns(data, required_cols, context="_build_monthly_status_table")

    data = _prepare_date_parts(data, date_column)
    data = _apply_filters(data, filters)

    data = data[
        data[value_column].notna() &
        data[status_column].notna()
    ].copy()

    data = data[data[status_column].isin(statuses)].copy()

    result = (
        data.pivot_table(
            index=["year", "month_number", "month"],
            columns=status_column,
            values=value_column,
            aggfunc="count",
            fill_value=0
        )
        .reset_index()
    )

    for status in statuses:
        if status not in result.columns:
            result[status] = 0

    result["Total general"] = result[statuses].sum(axis=1)

    ordered_cols = ["year", "month_number", "month"] + statuses + ["Total general"]
    result = result[ordered_cols].sort_values(["year", "month_number"]).reset_index(drop=True)

    return result


def add_calculated_columns(df: pd.DataFrame) -> pd.DataFrame:
    data = df.copy()

    required_cols = [
        "full_control_starting_date",
        "renewal_date",
        "full_control_ending_date",
        "subscription_id",
        "renewal_number",
    ]
    _validate_required_columns(data, required_cols, context="add_calculated_columns")

    data["full_control_starting_date"] = _normalize_date_column(data, "full_control_starting_date")
    data["renewal_date"] = _normalize_date_column(data, "renewal_date")
    data["full_control_ending_date"] = _normalize_date_column(data, "full_control_ending_date")
    data["renewal_number"] = pd.to_numeric(data["renewal_number"], errors="coerce")

    data["Unique Subscription Flag"] = (~data["subscription_id"].duplicated()).astype(int)

    data["Is Reactivation Renewal"] = (
        (
            data["full_control_starting_date"].dt.normalize()
            == data["renewal_date"].dt.normalize()
        )
        & (data["renewal_number"] == 1)
    ).astype(int)

    react_renewal_1_ids = set(
        data.loc[data["Is Reactivation Renewal"] == 1, "subscription_id"]
        .dropna()
        .astype(str)
    )

    data["Reactiv renewal 1"] = data["subscription_id"].where(
        data["subscription_id"].astype(str).isin(react_renewal_1_ids),
        pd.NA
    )

    return data


# Customers that joined the program (enrollment date), does not matter how (email or cancelation flow)
def customers_joined_program(df: pd.DataFrame) -> pd.DataFrame:
    return _build_monthly_count_table(
        df=df,
        date_column="full_control_starting_date",
        value_column="subscription_id",
        output_column_name="count_subscription_id",
        filters={"Unique Subscription Flag": 1},
    )


# Customers that joined + bought same day EMAIL
def reactivation_renewals(df: pd.DataFrame) -> pd.DataFrame:
    return _build_monthly_count_table(
        df=df,
        date_column="renewal_date",
        value_column="Is Reactivation Renewal",
        output_column_name="count_is_reactivation_renewal",
        filters={"Is Reactivation Renewal": 1},
    )


# Customers that joined + bought same day EMAIL more than once
def enrolled_bought_same_day_and_bought_more_than_once(df: pd.DataFrame) -> pd.DataFrame:
    return _build_monthly_count_table(
        df=df,
        date_column="full_control_starting_date",
        value_column="subscription_id",
        output_column_name="count_subscription_id",
        filters={
            "renewal_number": 2,
            "Reactiv renewal 1": {"op": "notna"},
        },
    )


# Customers that joined + bought later on (different date vs. join)
def first_renewal_not_reactivation(df: pd.DataFrame) -> pd.DataFrame:
    return _build_monthly_count_table(
        df=df,
        date_column="full_control_starting_date",
        value_column="subscription_id",
        output_column_name="count_subscription_id",
        filters={
            "renewal_number": 1,
            "Is Reactivation Renewal": 0,
        },
    )


# Customers joined that never bought
def no_renewals_yet(df: pd.DataFrame) -> pd.DataFrame:
    return _build_monthly_count_table(
        df=df,
        date_column="full_control_starting_date",
        value_column="subscription_id",
        output_column_name="count_subscription_id",
        filters={
            "renewal_number": 0,
        },
    )


# active (cohort) (active + processing)
def unique_subscriptions_active_processing(df: pd.DataFrame) -> pd.DataFrame:
    return _build_monthly_status_table(
        df=df,
        date_column="full_control_starting_date",
        value_column="subscription_id",
        status_column="subscription_status",
        statuses=["ACTIVE", "PROCESSING"],
        filters={
            "Unique Subscription Flag": 1,
        },
    )


# active (cohort) (active + processing + on hold)
def unique_subscriptions_active_processing_onhold(df: pd.DataFrame) -> pd.DataFrame:
    return _build_monthly_status_table(
        df=df,
        date_column="full_control_starting_date",
        value_column="subscription_id",
        status_column="subscription_status",
        statuses=["ACTIVE", "PROCESSING", "ON_HOLD"],
        filters={
            "Unique Subscription Flag": 1,
        },
    )


# Program Renewals (includes offer 1)
def all_renewals_after_enrollment(df: pd.DataFrame) -> pd.DataFrame:
    return _build_monthly_count_table(
        df=df,
        date_column="renewal_date",
        value_column="subscription_id",
        output_column_name="count_subscription_id",
        filters={
            "renewal_number": {"op": "ne", "value": 0},
        },
    )


# Renewals that correspond to customers that joined FC and had at least 2 renewals since then
def second_renewal(df: pd.DataFrame) -> pd.DataFrame:
    return _build_monthly_count_table(
        df=df,
        date_column="renewal_date",
        value_column="subscription_id",
        output_column_name="count_subscription_id",
        filters={
            "renewal_number": 2,
        },
    )


# Renewals 2+
def second_or_more_renewals(df: pd.DataFrame) -> pd.DataFrame:
    return _build_monthly_count_table(
        df=df,
        date_column="renewal_date",
        value_column="subscription_id",
        output_column_name="count_subscription_id",
        filters={
            "renewal_number": {"op": "gt", "value": 1},
        },
    )


def build_full_control_tables(prepared_df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """
    Genera todas las tablas necesarias para llenar el tracker.
    """
    return {
        "customers_joined_program": customers_joined_program(prepared_df),
        "reactivation_renewals": reactivation_renewals(prepared_df),
        "enrolled_bought_same_day_and_bought_more_than_once": enrolled_bought_same_day_and_bought_more_than_once(prepared_df),
        "first_renewal_not_reactivation": first_renewal_not_reactivation(prepared_df),
        "no_renewals_yet": no_renewals_yet(prepared_df),
        "unique_subscriptions_active_processing": unique_subscriptions_active_processing(prepared_df),
        "unique_subscriptions_active_processing_onhold": unique_subscriptions_active_processing_onhold(prepared_df),
        "all_renewals_after_enrollment": all_renewals_after_enrollment(prepared_df),
        "second_renewal": second_renewal(prepared_df),
        "second_or_more_renewals": second_or_more_renewals(prepared_df),
    }


def _normalize_month_header_value(value) -> pd.Timestamp:
    """
    Convierte el valor del header del template (datetime o texto) a Timestamp
    apuntando al primer día del mes.
    """
    if value is None:
        return pd.NaT

    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        raise ValueError(f"No se pudo interpretar el mes del header del template: {value}")

    return pd.Timestamp(year=parsed.year, month=parsed.month, day=1)


def _get_template_months(ws) -> list[tuple[int, int, int]]:
    """
    Lee los meses del template en B:Z (row 2) y devuelve:
    [(col_idx, year, month), ...]
    """
    months = []

    for col_idx in range(TEMPLATE_FIRST_MONTH_COL, TEMPLATE_LAST_MONTH_COL + 1):
        header_value = ws.cell(TEMPLATE_MONTH_HEADER_ROW, col_idx).value
        month_ts = _normalize_month_header_value(header_value)
        months.append((col_idx, month_ts.year, month_ts.month))

    return months


def _get_table_value_column(table_df: pd.DataFrame) -> str:
    """
    Determina qué columna usar para escribir en el template.
    - Si existe 'Total general', usa esa (para tablas por status)
    - Si no, usa la única columna distinta de year/month_number/month
    """
    excluded = {"year", "month_number", "month"}

    if "Total general" in table_df.columns:
        return "Total general"

    value_cols = [col for col in table_df.columns if col not in excluded]
    if len(value_cols) != 1:
        raise ValueError(
            f"No se pudo determinar la columna de valores. Columnas disponibles: {table_df.columns.tolist()}"
        )

    return value_cols[0]


def _table_to_month_value_map(table_df: pd.DataFrame) -> dict[tuple[int, int], float]:
    """
    Convierte una tabla mensual a un diccionario:
    {(year, month): value}
    """
    data = table_df.copy()
    value_col = _get_table_value_column(data)

    _validate_required_columns(
        data,
        ["year", "month_number", value_col],
        context="_table_to_month_value_map"
    )

    data["year"] = pd.to_numeric(data["year"], errors="coerce").astype("Int64")
    data["month_number"] = pd.to_numeric(data["month_number"], errors="coerce").astype("Int64")
    data[value_col] = pd.to_numeric(data[value_col], errors="coerce").fillna(0)

    month_value_map = {}
    for _, row in data.iterrows():
        key = (int(row["year"]), int(row["month_number"]))
        month_value_map[key] = float(row[value_col])

    return month_value_map


def fill_full_control_tracker(
    template_path: str,
    output_path: str,
    prepared_df: pd.DataFrame,
    sheet_name: str = FULL_CONTROL_SHEET_NAME,
) -> str:
    """
    Llena el archivo tracker en la hoja 'Full control'.

    Meses:
    - Lee automáticamente los headers de B:Z en la fila 2
    - Si un mes no existe en la tabla, escribe 0

    Importante:
    - En el template recibido, la fila de 'active + processing + on hold' es la 26
      (la 25 es porcentaje)
    - En el template recibido, la fila 42 es 'second_renewal' y la 43 es 'second_or_more_renewals'
    """
    tables = build_full_control_tables(prepared_df)

    wb = load_workbook(template_path)
    ws = wb[sheet_name]

    # Leer meses del template
    template_months = _get_template_months(ws)

    # Row mapping basado en los labels reales del template
    row_mapping = {
        "customers_joined_program": 4,
        "reactivation_renewals": 15,
        "enrolled_bought_same_day_and_bought_more_than_once": 16,
        "first_renewal_not_reactivation": 20,
        "no_renewals_yet": 21,
        "unique_subscriptions_active_processing": 24,
        "unique_subscriptions_active_processing_onhold": 26,   
        "all_renewals_after_enrollment": 40,
        "second_renewal": 42,                                  
        "second_or_more_renewals": 43,                        
    }

    for table_name, row_idx in row_mapping.items():
        month_value_map = _table_to_month_value_map(tables[table_name])

        for col_idx, year, month in template_months:
            value = month_value_map.get((year, month), 0)
            ws.cell(row=row_idx, column=col_idx).value = float(value)

    # Forzar recálculo de fórmulas al abrir en Excel
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    TEMPLATE_PATH = "Ecomm initiatives trackers.xlsx"
    OUTPUT_PATH = "Ecomm initiatives trackers - filled.xlsx"

    raw_df = execute_query(fc_query)
    prepared_df = add_calculated_columns(raw_df)

    # Tablas individuales (opcional)
    ta_result = customers_joined_program(prepared_df)
    tb_result = reactivation_renewals(prepared_df)
    tc_result = enrolled_bought_same_day_and_bought_more_than_once(prepared_df)
    td_result = first_renewal_not_reactivation(prepared_df)
    te_result = no_renewals_yet(prepared_df)
    tf_result = unique_subscriptions_active_processing(prepared_df)
    tg_result = unique_subscriptions_active_processing_onhold(prepared_df)
    th_result = all_renewals_after_enrollment(prepared_df)
    ti_result = second_renewal(prepared_df)
    tj_result = second_or_more_renewals(prepared_df)

    print("TA")
    print(ta_result)

    print("\nTB")
    print(tb_result)

    print("\nTC")
    print(tc_result)

    print("\nTD")
    print(td_result)

    print("\nTE")
    print(te_result)

    print("\nTF")
    print(tf_result)

    print("\nTG")
    print(tg_result)

    print("\nTH")
    print(th_result)

    print("\nTI")
    print(ti_result)

    print("\nTJ")
    print(tj_result)

    output_file = fill_full_control_tracker(
        template_path=TEMPLATE_PATH,
        output_path=OUTPUT_PATH,
        prepared_df=prepared_df,
    )

    print(f"\nArchivo generado: {output_file}")
