import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def normalize_admin_status(s) -> str:
    if pd.isna(s):
        return ""
    return str(s).strip().upper()


def is_verified_admin_status(admin_status: str) -> bool:
    """
    Una review es verificada si adminStatus contiene 'VERIFIED'
    (case-insensitive). Ej: 'Verified Review'
    """
    return "VERIFIED" in normalize_admin_status(admin_status)


def safe_int(v):
    try:
        if pd.isna(v):
            return None
        return int(v)
    except:
        return None


def ajustar_ancho_columnas(archivo_excel: str) -> None:
    """
    Ajusta automáticamente el ancho de las columnas del Excel
    (misma lógica que usas en otros reportes)
    """
    try:
        wb = load_workbook(archivo_excel)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if cell.value not in (None, ""):
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

                ws.column_dimensions[column_letter].width = min(max_length + 2, 55)

        wb.save(archivo_excel)
        print(f"✓ Ancho de columnas ajustado para {archivo_excel}")

    except Exception as e:
        print(f"✗ Error ajustando ancho de columnas: {e}")


def read_reviews_json(json_path: str) -> dict:
    with open(json_path, "r", encoding="utf-8") as f:
        return json.load(f)


def build_report_from_reviews(data: dict):
    """
    Filtros:
    - adminStatus contiene 'verified'
    - rating = 4 o 5
    """

    items = data.get("itemsList", []) or []
    records = []

    for item in items:
        fb = item.get("feedback", {}) or {}

        if not is_verified_admin_status(fb.get("adminStatus", "")):
            continue

        rating = safe_int(fb.get("rating"))
        if rating not in (4, 5):
            continue

        recomendation = 0
        if rating == 5:
            recomendation = 10
        if rating == 4:
            recomendation = 8

        records.append(
            {
                "Product": fb.get("productName") or item.get("title"),
                "SKU": fb.get("sku"),
                "Order ID": fb.get("orderId"),
                "headline": fb.get("title"),
                "comment": fb.get("feedBack"),
                "nickname": fb.get("nameOnAmazon"),
                "email": fb.get("email"),
                "recomendation": recomendation,
                "overall_rating": rating,
                "date": fb.get("createdAt"),
            }
        )

    table_df = pd.DataFrame(
        records,
        columns=[
            "Product",
            "SKU",
            "Order ID",
            "headline",
            "comment",
            "nickname",
            "email",
            "recomendation",
            "overall_rating",
            "date",
        ],
    )

    summary_df = pd.DataFrame(
        [
            {"Metric": "Total items in JSON", "Value": len(items)},
            {
                "Metric": "Verified reviews (adminStatus) with rating 4-5",
                "Value": len(table_df),
            },
        ]
    )

    return table_df, summary_df


def main():
    """
    Lee el archivo fijo 'reviews.json'
    Genera:
      - verified_reviews_4_5.xlsx
      - verified_reviews_4_5.csv
    """
    input_file = "reviews.json"
    print(f"Leyendo {input_file}...")

    data = read_reviews_json(input_file)

    print("Procesando reviews verificadas con rating 4-5...")
    table_df, summary_df = build_report_from_reviews(data)

    if len(table_df) == 0:
        print("⚠️ No se encontraron reviews que cumplan los criterios.")

    output_xlsx = "verified_reviews_4_5.xlsx"
    with pd.ExcelWriter(output_xlsx, engine="openpyxl", datetime_format="yyyy-mm-dd hh:mm:ss.000") as writer:
        table_df.to_excel(writer, sheet_name="Verified Reviews (4-5)", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

    ajustar_ancho_columnas(output_xlsx)

    output_csv = "verified_reviews_4_5.csv"
    table_df.to_csv(output_csv, index=False, encoding="utf-8-sig")

    print(f"\n✅ Reportes generados:")
    print(f" - {output_xlsx}")
    print(f" - {output_csv}")

    if len(table_df) > 0:
        print("\n📋 Preview:")
        print(table_df.head(10).to_string(index=False))


if __name__ == "__main__":
    main()
