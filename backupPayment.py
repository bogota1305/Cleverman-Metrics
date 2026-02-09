import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from modules.database_queries import execute_query


def safe_pct(n: int, d: int) -> float:
    return (n / d) if d else 0.0


def normalize_status(s) -> str:
    if pd.isna(s):
        return ""
    return str(s).strip().upper()


def normalize_backup_flag(v) -> bool:
    """
    metadata ->> '$.isBackupPayment' suele venir como:
      'true' / 'false' / NULL
    Convertimos a bool.
    """
    if pd.isna(v):
        return False
    s = str(v).strip().lower()
    return s in ("true", "1", "yes", "y", "t")


def ajustar_ancho_columnas(archivo_excel: str) -> None:
    """
    Ajusta automáticamente el ancho de las columnas en un archivo Excel
    (igual al enfoque que ya usas)
    """
    try:
        wb = load_workbook(archivo_excel)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if cell.value is not None and cell.value != "":
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass

                adjusted_width = min(max_length + 2, 55)
                ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(archivo_excel)
        print(f"✓ Ancho de columnas ajustado para {archivo_excel}")

    except Exception as e:
        print(f"✗ Error ajustando el ancho de columnas para {archivo_excel}: {e}")


def build_report_from_payments(df: pd.DataFrame):
    """
    Reglas:
    - Cada fila = intento
    - Un "pago" = entityId
    - Had Error = si existe >= 1 FAILED dentro del grupo
    - Resolved = si Had Error y el ÚLTIMO intento (por createdAt) es SUCCESS
    - Resolved by Backup = si Resolved y el ÚLTIMO intento SUCCESS tiene backupPayment = true
    """

    # Parse createdAt con milisegundos: '2025-12-12 16:06:21.419'
    df["createdAt"] = pd.to_datetime(df["createdAt"], errors="coerce")
    if df["createdAt"].isna().any():
        bad = df[df["createdAt"].isna()][["id", "entityId", "createdAt"]].head(10)
        raise ValueError(
            "Hay filas con 'createdAt' inválido. Ejemplos:\n"
            f"{bad.to_string(index=False)}"
        )

    df["_status_norm"] = df["status"].apply(normalize_status)
    df["_backup_bool"] = df["backupPayment"].apply(normalize_backup_flag)

    records = []
    df_sorted = df.sort_values("createdAt")

    for entity_id, g in df_sorted.groupby("entityId", sort=False):
        g = g.sort_values("createdAt")

        attempts = int(len(g))
        customer_id = g["customerId"].iloc[0] if "customerId" in g.columns else None

        had_error = (g["_status_norm"] == "FAILED").any()

        last_row = g.iloc[-1]
        last_status = last_row["_status_norm"]
        last_backup = bool(last_row["_backup_bool"])

        resolved = bool(had_error and last_status == "SUCCESS")
        resolved_by_backup = bool(resolved and last_backup)

        records.append(
            {
                "customerId": customer_id,
                "entityId": entity_id,
                "Attempts": attempts,
                "Had Error": "Yes" if had_error else "No",
                "Resolved": "Yes" if resolved else "No",
                "Resolved by Backup": "Yes" if resolved_by_backup else "No",
                "First Attempt UTC": g["createdAt"].min(),
                "Last Attempt UTC": g["createdAt"].max(),
                "Last Attempt Status": last_status,
                "Last Attempt BackupPayment": "true" if last_backup else "false",
                "Last Attempt ID": last_row["id"],
            }
        )

    detail_df = pd.DataFrame(records)

    total_payments = int(len(detail_df))
    total_errors = int((detail_df["Had Error"] == "Yes").sum())
    total_resolved = int((detail_df["Resolved"] == "Yes").sum())
    total_resolved_backup = int((detail_df["Resolved by Backup"] == "Yes").sum())

    # OJO: mantengo exactamente tus fórmulas pedidas (aunque las 2 primeras tasas quedan iguales)
    summary_df = pd.DataFrame(
        [
            {"Metric": "Total payments", "Value": total_payments},
            {"Metric": "Total payment errors (payments with ≥1 FAILED)", "Value": total_errors},
            {"Metric": "Total resolved payments (error resolved)", "Value": total_resolved},
            {"Metric": "Total resolved by backup", "Value": total_resolved_backup},
            {
                "Metric": "Porcentaje de errores de pago (total resolved / total errors)",
                "Value": safe_pct(total_resolved, total_errors),
            },
            {
                "Metric": "Porcentaje de pagos resueltos (total resolved / total errors)",
                "Value": safe_pct(total_resolved, total_errors),
            },
            {
                "Metric": "Porcentaje de resueltos por backup (resolved by backup / total resolved)",
                "Value": safe_pct(total_resolved_backup, total_resolved),
            },
        ]
    )

    backup_resolved_df = detail_df[detail_df["Resolved by Backup"] == "Yes"].copy()
    return detail_df, backup_resolved_df, summary_df


def main(startDate: str, endDate: str):
    """
    startDate / endDate ejemplos:
      '2025-12-01 00:00:00'
      '2026-01-01 00:00:00'

    Si prefieres solo fechas:
      '2025-12-01'
    MySQL suele castear bien, pero lo más seguro es datetime completo.
    """

    query = f"""
    SELECT
        id,
        entityId,
        createdAt,
        customerId,
        status,
        metadata ->> '$.isBackupPayment' as backupPayment
    FROM prod_sales_and_subscriptions.payments
    WHERE createdAt >= '{startDate}'
      AND createdAt < '{endDate}'
    ;
    """

    print("Ejecutando consulta de payments...")
    df = execute_query(query)

    if df is None or len(df) == 0:
        print("⚠️ La consulta no devolvió filas para ese rango de fechas.")
        return

    # Report
    print("Procesando pagos agrupados por entityId...")
    detail_df, backup_resolved_df, summary_df = build_report_from_payments(df)

    # Guardar Excel (mismas hojas que el otro análisis)
    nombre_archivo = f"backup_payment_methods_{startDate}_to_{endDate}.xlsx".replace(":", "-").replace(" ", "_")

    with pd.ExcelWriter(nombre_archivo, engine="openpyxl", datetime_format="yyyy-mm-dd hh:mm:ss.000") as writer:
        backup_resolved_df.to_excel(writer, sheet_name="Backup Resolved", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # Si un día quieres el detalle completo, descomenta:
        # detail_df.to_excel(writer, sheet_name="All Payments (Detail)", index=False)

    ajustar_ancho_columnas(nombre_archivo)

    # Prints de control
    total_payments = len(detail_df)
    total_errors = (detail_df["Had Error"] == "Yes").sum()
    total_resolved = (detail_df["Resolved"] == "Yes").sum()
    total_resolved_backup = (detail_df["Resolved by Backup"] == "Yes").sum()

    print(f"\n✅ Reporte generado: {nombre_archivo}")
    print(f"Total payments: {total_payments}")
    print(f"Total payment errors: {total_errors}")
    print(f"Total resolved payments: {total_resolved}")
    print(f"Total resolved by backup: {total_resolved_backup}")

    print("\n📋 Top 10 entityId resueltos por backup:")
    if len(backup_resolved_df) > 0:
        print(backup_resolved_df[["entityId", "customerId", "Attempts", "Last Attempt UTC"]].head(10))
    else:
        print("No hubo pagos resueltos por backup en este rango.")


if __name__ == "__main__":
    # Ajusta el rango a lo que necesites
    main("2025-01-01 00:00:00", "2026-01-01 00:00:00")
